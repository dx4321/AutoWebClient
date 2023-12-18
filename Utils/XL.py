from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

from Utils.logger import logger
from Utils.read_config import AppConfig, read_yaml_config_and_convert_to_app_config


class XL:
    def __init__(self, _app_config: AppConfig):
        self.file_name = _app_config.team_city.history_file_path
        self._app_config = _app_config
        try:
            wb = load_workbook(self.file_name)
            ws1 = wb.active

        except Exception as _e:
            self.create_first_data()
            logger.debug(f"Создали файл - {_app_config.team_city.history_file_path}")

    def create_first_data(self):
        """ Если нет файла с историями, то создать его, заполнить заголовки теми что в конфиге (RanorexDeployments) """

        wb = Workbook()
        ws = wb.active
        programs = [*self._app_config.team_city.for_test_programs_info]
        for i, value in enumerate(programs):
            ws.cell(1, i + 1, value)

        wb.save(self._app_config.team_city.history_file_path)

    def save_to_last_line(self, soft_name: str, last_str_val):
        """ Найти столбец с нужной программой и сохранить в самом конце столбца необходимую строку """

        wb = load_workbook(self.file_name)  # Прочитываем книгу
        ws = wb.active  # Получаем значения с активного листа
        firs_row = ws[1]  # 1-ая строка

        required_column = None  # Указываем искомый столбец с программой

        # получаем индекс столбца с нужной программой
        for i, val in enumerate(firs_row):
            if val.value == soft_name:
                required_column = i + 1

        word = get_column_letter(required_column)  # Получить букву столбца
        stolbec = ws[word]  # Получить объект со всеми значениями столбца

        # Получаем все не пустые значения в столбце
        list_vals_in_stolbec = [lat.value for lat in stolbec if lat.value]

        # Получаем кол-во значений в столбце
        last_row_in_column = len(list_vals_in_stolbec)

        # По координатам ячейки сохраняем значение ("A:6")
        ws[
            str(
                get_column_letter(required_column) + str(last_row_in_column + 1)
            )] = last_str_val

        # Сохраняем рабочую книгу
        wb.save(self.file_name)

    def get_last_str_for_soft(self, soft_name):
        """ Получить последнюю строку по софту """

        wb = load_workbook(self.file_name)  # Прочитываем книгу
        ws = wb.active  # Получаем значения с активного листа
        firs_row = ws[1]  # 1-ая строка

        required_column = None  # Указываем искомый столбец с программой

        # получаем индекс столбца с нужной программой
        for i, val in enumerate(firs_row):
            if val.value == soft_name:
                required_column = i + 1

        word = get_column_letter(required_column)  # Получить букву столбца
        stolbec = ws[word]  # Получить объект со всеми значениями столбца

        # Получаем все не пустые значения в столбце
        list_vals_in_stolbec = [lat.value for lat in stolbec if lat.value]

        # Получаем кол-во значений в столбце
        last_row_in_column = len(list_vals_in_stolbec)

        # По координатам ячейки сохраняем значение ("A:6")
        last_str_val = ws[
            str(
                get_column_letter(required_column) + str(last_row_in_column)
            )]

        return last_str_val.value


# Tests
if __name__ == '__main__':
    app_config = read_yaml_config_and_convert_to_app_config("../config.yaml")
    apps = app_config.team_city.for_test_programs_info

    xl = XL(app_config)
    data = [*app_config.team_city.for_test_programs_info.keys()]
    xl.save_to_last_line(data[0], "test")
    get_last_soft_data = xl.get_last_str_for_soft(data[1])

    logger.debug(get_last_soft_data)
